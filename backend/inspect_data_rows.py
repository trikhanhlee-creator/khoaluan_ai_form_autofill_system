#!/usr/bin/env python3
"""
Detailed inspection of data rows
"""
import openpyxl

file_path = r"c:\Users\KHANH\Downloads\demoexcel.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("\n" + "="*80)
print("DETAILED ROW INSPECTION")
print("="*80)

print("\nRows 1-5 (first 5 rows):")
for row_idx in range(1, 6):
    row_data = [ws.cell(row_idx, c).value for c in range(1, 10)]
    print(f"\nRow {row_idx}: {row_data}")

print("\n" + "="*80)
print("ANALYSIS:")
print("="*80)

# Check if Row 3 is a data row or label row
row3_vals = [ws.cell(3, c).value for c in range(1, 10)]
print(f"\nRow 3 values: {row3_vals}")

# Check if all values are percentage labels
is_percentage_row = all(
    (val is None or str(val).strip() == '' or '%' in str(val).lower())
    for val in row3_vals
)
print(f"Is Row 3 a percentage label row? {is_percentage_row}")

# Check Row 4 (first data row)
row4_vals = [ws.cell(4, c).value for c in range(1, 10)]
print(f"\nRow 4 values: {row4_vals}")

is_data_row = any(
    (val is not None and str(val).strip() and '%' not in str(val).lower())
    for val in row4_vals
)
print(f"Is Row 4 a data row? {is_data_row}")

print("\n" + "="*80 + "\n")

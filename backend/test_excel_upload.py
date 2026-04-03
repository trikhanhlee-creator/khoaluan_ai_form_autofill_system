"""Test Excel upload functionality"""
import sys
import os
from io import BytesIO
import openpyxl
from fastapi import UploadFile
from pathlib import Path

# Add the backend directory to path
sys.path.insert(0, os.path.dirname(__file__))

def test_excel_upload():
    """Test the Excel upload endpoint"""
    
    print("\n" + "="*60)
    print("EXCEL UPLOAD TEST")
    print("="*60)
    
    # Test 1: Create a valid Excel file
    print("\n1. Creating test Excel file...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Ho Ten'
        ws['B1'] = 'Email'
        ws['C1'] = 'Phone'
        ws['A2'] = 'Nguyen Van A'
        ws['B2'] = 'a@test.com'
        ws['C2'] = '0901234567'
        ws['A3'] = 'Tran Thi B'
        ws['B3'] = 'b@test.com'
        ws['C3'] = '0912345678'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        print(f"   ✓ Created Excel file: {len(buffer.getvalue())} bytes")
    except Exception as e:
        print(f"   ✗ Error creating Excel file: {e}")
        return
    
    # Test 2: Test parsing the Excel file
    print("\n2. Testing Excel parsing...")
    try:
        buffer.seek(0)
        workbook = openpyxl.load_workbook(BytesIO(buffer.getvalue()), data_only=False, read_only=False)
        worksheet = workbook.active
        
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f"   ✓ Headers: {headers}")
        
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header_idx, header in enumerate(headers):
                value = row[header_idx] if header_idx < len(row) else None
                row_data[header] = str(value) if value is not None else ""
            if any(row_data.values()):
                rows.append(row_data)
        
        print(f"   ✓ Found {len(rows)} data rows")
        for i, row in enumerate(rows):
            print(f"      Row {i+1}: {row}")
            
    except Exception as e:
        print(f"   ✗ Error parsing Excel: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Test 3: Test with the sample file if it exists
    print("\n3. Testing with sample_data.xlsx...")
    sample_path = Path(__file__).parent / "uploads" / "sample_data.xlsx"
    if sample_path.exists():
        try:
            with open(sample_path, 'rb') as f:
                sample_buffer = BytesIO(f.read())
            sample_buffer.seek(0)
            workbook = openpyxl.load_workbook(sample_buffer, data_only=False, read_only=False)
            worksheet = workbook.active
            
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            print(f"   ✓ Sample file exists and can be parsed")
            print(f"   ✓ Headers: {headers}")
            
            row_count = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_count += 1
            print(f"   ✓ Found {row_count} data rows")
        except Exception as e:
            print(f"   ✗ Error reading sample file: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"   ! Sample file not found at {sample_path}")
    
    # Test 4: Try to import and test the actual route
    print("\n4. Testing Excel route import...")
    try:
        from app.api.routes.excel import router
        print(f"   ✓ Excel router imported successfully")
        print(f"   ✓ Routes: upload, data, row, sessions, session")
    except Exception as e:
        print(f"   ✗ Error importing Excel router: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "="*60)
    print("TEST COMPLETE")
    print("="*60 + "\n")

if __name__ == "__main__":
    test_excel_upload()

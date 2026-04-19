"""Test Excel upload via HTTP API"""
import os
import sys
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import requests
import openpyxl
from io import BytesIO

def test_excel_upload_http():
    """Test Excel upload via HTTP"""
    
    print("\n" + "="*70)
    print("TESTING EXCEL UPLOAD VIA HTTP API")
    print("="*70 + "\n")
    
    # Create test Excel file
    print("1. Creating test Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Ten'
    ws['B1'] = 'Email'
    ws['A2'] = 'Test User'
    ws['B2'] = 'test@email.com'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    print(f"   ✓ Excel file created: {len(buffer.getvalue())} bytes\n")
    
    # Test upload 1: Using test file
    print("2. Testing upload with test file...")
    files = {'file': ('test.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
    print(f"   Status: {response.status_code}")
    print(f"   Response: {response.text}")
    if response.status_code == 200:
        print(f"   ✓ SUCCESS\n")
    else:
        print(f"   ✗ ERROR\n")
    
    # Test upload 2: Using sample file
    print("3. Testing upload with sample_data.xlsx...")
    sample_path = BACKEND_DIR / "uploads" / "sample_data.xlsx"
    if sample_path.exists():
        try:
            with open(sample_path, 'rb') as f:
                files = {'file': ('sample_data.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
            print(f"   Status: {response.status_code}")
            print(f"   Response: {response.text}")
            if response.status_code == 200:
                print(f"   ✓ SUCCESS\n")
            else:
                print(f"   ✗ ERROR\n")
        except Exception as e:
            print(f"   ✗ Error: {e}\n")
    else:
        print(f"   ! Sample file not found\n")
    
    # Test list sessions
    print("4. Testing list sessions...")
    response = requests.get('http://127.0.0.1:8000/api/excel/sessions')
    print(f"   Status: {response.status_code}")
    print(f"   Response: {response.text}\n")
    
    print("=" * 70 + "\n")

if __name__ == "__main__":
    try:
        test_excel_upload_http()
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()

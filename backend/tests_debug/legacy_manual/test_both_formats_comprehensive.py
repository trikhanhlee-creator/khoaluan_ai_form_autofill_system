#!/usr/bin/env python3
"""
Comprehensive test: Both .xlsx and .xls formats with various scenarios
Ensures both fixes work correctly without interfering with each other
"""
import requests
import time
import xlwt
from openpyxl import Workbook
import os

BASE_URL = "http://localhost:8000"

def create_test_files():
    """Create test files for both formats"""
    
    print("   Creating .xlsx files...")
    # XLSX with empty first row
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, None])  # Empty
    ws.append(['Name', 'Email', 'Phone'])
    ws.append(['Person A', 'a@test.com', '0123456789'])
    wb.save('test_xlsx_empty_row.xlsx')
    
    # XLSX normal
    wb = Workbook()
    ws = wb.active
    ws.append(['Customer', 'Contact', 'City'])
    ws.append(['ABC Corp', 'contact@abc.com', 'Hanoi'])
    ws.append(['XYZ Ltd', 'info@xyz.com', 'HCMC'])
    wb.save('test_xlsx_normal.xlsx')
    
    print("   Creating .xls files...")
    # XLS with empty first row
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Data')
    # Row 0 empty
    sheet.write(1, 0, 'Product')
    sheet.write(1, 1, 'Price')
    sheet.write(1, 2, 'Stock')
    sheet.write(2, 0, 'Laptop')
    sheet.write(2, 1, 15000000)
    sheet.write(2, 2, 5)
    workbook.save('test_xls_empty_row.xls')
    
    # XLS normal
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Data')
    sheet.write(0, 0, 'Department')
    sheet.write(0, 1, 'Head')
    sheet.write(0, 2, 'Budget')
    sheet.write(1, 0, 'IT')
    sheet.write(1, 1, 'Tran Van A')
    sheet.write(1, 2, 5000000000)
    sheet.write(2, 0, 'HR')
    sheet.write(2, 1, 'Nguyen Thi B')
    sheet.write(2, 2, 2000000000)
    workbook.save('test_xls_normal.xls')

def test_file(filename, file_type):
    """Test a single file"""
    print(f"\n  📁 {filename} ({file_type})")
    
    try:
        with open(filename, 'rb') as f:
            files = {'file': (filename, f)}
            response = requests.post(f"{BASE_URL}/api/excel/upload", files=files, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            print(f"     ✅ Status {response.status_code}")
            print(f"        Headers: {len(data.get('headers', []))} columns")
            print(f"        Data rows: {data.get('total_rows', 0)}")
            return True
        else:
            error = response.text[:100]
            print(f"     ❌ Status {response.status_code}")
            print(f"        Error: {error}")
            return False
    except Exception as e:
        print(f"     ❌ Exception: {str(e)[:80]}")
        return False

def main():
    print("\n" + "=" * 70)
    print("✅ COMPREHENSIVE TEST: .xlsx + .xls with improved parsing")
    print("=" * 70)
    
    # Wait for server
    print("\nWaiting for server...")
    for i in range(15):
        try:
            requests.get(f"{BASE_URL}/health", timeout=1)
            print("✅ Server ready!\n")
            break
        except:
            time.sleep(1)
            if i == 14:
                print("❌ Server not responding")
                return False
    
    # Create files
    print("Creating test files...")
    create_test_files()
    print("✅ Test files created\n")
    
    # Test scenarios
    print("=" * 70)
    print("🧪 TESTING BOTH FORMATS")
    print("=" * 70)
    
    test_cases = [
        ('test_xlsx_empty_row.xlsx', '.xlsx (empty first row)'),
        ('test_xlsx_normal.xlsx', '.xlsx (normal)'),
        ('test_xls_empty_row.xls', '.xls (empty first row)'),
        ('test_xls_normal.xls', '.xls (normal)'),
    ]
    
    results = {}
    for filename, desc in test_cases:
        results[filename] = test_file(filename, desc)
    
    # Summary
    print("\n" + "=" * 70)
    print("📊 FINAL RESULTS")
    print("=" * 70)
    
    passed = sum(1 for v in results.values() if v)
    total = len(results)
    
    print(f"\n✓ Passed: {passed}/{total} tests")
    
    for filename, success in results.items():
        status = "✅" if success else "❌"
        print(f"  {status} {filename}")
    
    if passed == total:
        print("\n" + "=" * 70)
        print("🎉 SUCCESS: Both .xlsx and .xls formats working perfectly!")
        print("=" * 70)
        print("\nKey achievements:")
        print("  ✓ .xlsx files with empty first row now work")
        print("  ✓ .xlsx files with normal headers still work")
        print("  ✓ .xls files with empty first row now work")
        print("  ✓ .xls files with normal headers still work")
        print("  ✓ No conflicts between format handlers")
        return True
    else:
        print(f"\n⚠️  {total - passed} test(s) failed")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)

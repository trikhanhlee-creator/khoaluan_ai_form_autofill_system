#!/usr/bin/env python3
"""
Test XLSX upload with improved parsing
Tests both normal format and problematic formats
"""
import requests
import time
from openpyxl import Workbook
import os

BASE_URL = "http://localhost:8000"

def create_test_xlsx_files():
    """Create various XLSX test files"""
    
    files_created = {}
    
    # Test 1: Normal XLSX with headers in row 1
    print("   • Creating: normal_headers.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(['Ho Ten', 'Email', 'So Dien Thoai', 'Dia Chi'])
    ws.append(['Khanh Nguyen', 'khanh@company.com', '0987654321', 'Ha Noi'])
    ws.append(['Linh Tran', 'linh@company.com', '0912345678', 'Ho Chi Minh'])
    wb.save('normal_headers.xlsx')
    files_created['normal_headers.xlsx'] = {
        'name': 'Scenario 1: Normal XLSX (headers row 1)',
        'description': 'Standard XLSX with headers in first row'
    }
    
    # Test 2: XLSX with empty first row
    print("   • Creating: empty_first_row.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append([None, None, None, None])  # Empty row
    ws.append(['Tên Sinh Viên', 'MSSV', 'Lớp', 'GPA'])
    ws.append(['Trần Thị B', 'SV001', 'D21CQCN01', 3.5])
    ws.append(['Nguyễn Văn C', 'SV002', 'D21CQCN02', 3.8])
    wb.save('empty_first_row.xlsx')
    files_created['empty_first_row.xlsx'] = {
        'name': 'Scenario 2: XLSX with empty first row',
        'description': 'Headers in row 2 (row 1 is empty)'
    }
    
    # Test 3: XLSX with mixed data types
    print("   • Creating: mixed_data_types.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(['Tên KH', 'Tuổi', 'Năm Sinh', 'Điểm'])
    ws.append(['Phạm Van A', 28, 1996, 85.5])
    ws.append(['Lê Văn B', 25, 1999, 90.0])
    ws.append(['Đỗ Thị C', 30, 1994, 78.75])
    wb.save('mixed_data_types.xlsx')
    files_created['mixed_data_types.xlsx'] = {
        'name': 'Scenario 3: XLSX with mixed data types',
        'description': 'Text and numbers mixed, Vietnamese characters'
    }
    
    # Test 4: XLSX with empty rows in middle
    print("   • Creating: gaps_between_data.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(['Mã SP', 'Tên Sản Phẩm', 'Giá', 'Kho'])
    ws.append(['P001', 'Laptop', 15000000, 5])
    ws.append([None, None, None, None])  # Empty row
    ws.append(['P002', 'Chuột', 500000, 20])
    ws.append(['P003', 'Bàn Phím', 1000000, 15])
    wb.save('gaps_between_data.xlsx')
    files_created['gaps_between_data.xlsx'] = {
        'name': 'Scenario 4: XLSX with gaps',
        'description': 'Data rows separated by empty rows'
    }
    
    return files_created

def test_upload(filename, details):
    """Test uploading XLSX file"""
    
    scenario_name = details['name']
    description = details['description']
    
    print(f"\n  📋 {scenario_name}")
    print(f"     {description}")
    print(f"     File: {filename}")
    print("     " + "-" * 56)
    
    try:
        with open(filename, 'rb') as f:
            files = {'file': (filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{BASE_URL}/api/excel/upload", files=files, timeout=5)
        
        print(f"     HTTP Status: {response.status_code}")
        
        if response.status_code == 200:
            data = response.json()
            headers = data.get('headers', [])
            total_rows = data.get('total_rows', 0)
            message = data.get('message', '')
            
            print(f"     ✅ SUCCESS!")
            print(f"        • Headers: {headers}")
            print(f"        • Data rows: {total_rows}")
            print(f"        • Message: {message}")
            return True
        else:
            error_text = response.text[:150]
            print(f"     ❌ FAILED!")
            print(f"        Error: {error_text}")
            return False
    
    except Exception as e:
        print(f"     ❌ EXCEPTION!")
        print(f"        Error: {str(e)}")
        return False

def main():
    print("\n" + "=" * 70)
    print("🔧 XLSX PARSING FIX VERIFICATION TEST")
    print("Testing improved XLSX header detection")
    print("=" * 70)
    
    # Wait for server
    print("\nWaiting for server...")
    for i in range(15):
        try:
            requests.get(f"{BASE_URL}/health", timeout=1)
            print("✅ Server ready!\n")
            break
        except:
            time.sleep(1)
            if i == 14:
                print("❌ Server not responding")
                return False
    
    # Create test files
    print("Creating test XLSX files...")
    test_files = create_test_xlsx_files()
    print(f"✅ Created {len(test_files)} test files\n")
    
    # Run tests
    print("=" * 70)
    print("🧪 TESTING XLSX SCENARIOS")
    print("=" * 70)
    
    results = {}
    for filename, details in test_files.items():
        results[filename] = test_upload(filename, details)
    
    # Summary
    print("\n" + "=" * 70)
    print("📊 TEST SUMMARY")
    print("=" * 70)
    
    passed = sum(1 for v in results.values() if v)
    total = len(results)
    
    print(f"\nResults: {passed}/{total} tests passed")
    for filename, success in results.items():
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"  {status}: {filename}")
    
    if passed == total:
        print("\n✅ ALL XLSX TESTS PASSED!")
        print("\n🎉 XLSX fix verified: Can now handle empty first rows!")
        return True
    else:
        print(f"\n⚠️  Some tests failed ({total - passed} failures)")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)

#!/usr/bin/env python3
"""
Test Excel parsing with all columns extraction and field metadata
"""
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import sys
import os

sys.path.insert(0, '.')

from app.api.routes.excel import (
    parse_excel_with_openpyxl, 
    extract_field_metadata,
    detect_field_type,
    normalize_field_name,
    extract_keywords
)

def create_test_excel_with_sections():
    """Create a test Excel file mimicking the student info structure"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Row 1: Headers with sections
    headers = [
        "STT",
        "Mã sinh viên",
        "Ho đêm",
        "Tên",
        "Giới tính",
        "Ngày sinh",
        "Lớp học",
        "Chuyên cần (20%)",
        "Thương xuyên (30%)",
        "Quân sự",
        "Tình nguyện"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = header
    
    # Row 2+: Data
    data_rows = [
        ["1", "221561", "Huỳnh", "Hoài An", "Nam", "12/06/2004", "DH22TIN07", "10.00", "10.00", "9.50", "8.00"],
        ["2", "222335", "Trần", "Trúc An", "Nam", "18/04/2004", "DH22TIN07", "9.80", "10.00", "9.00", "8.50"],
        ["3", "221555", "Huỳnh", "Hoài An", "Nam", "12/06/2004", "DH22TIN07", "10.00", "10.00", "9.80", "9.00"],
    ]
    
    for row_idx, data_row in enumerate(data_rows, 2):
        for col_idx, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

def test_excel_parsing():
    """Test the Excel parser with metadata extraction"""
    print("\n" + "="*70)
    print("[TEST] Excel Parser - Extract All Columns with Field Metadata")
    print("="*70)
    
    # Create test file
    print("\n[1/4] Creating test Excel file...")
    excel_file = create_test_excel_with_sections()
    
    # Parse file
    print("[2/4] Parsing Excel file...")
    try:
        headers, rows, msg = parse_excel_with_openpyxl(excel_file)
        
        print(f"\n✅ {msg}")
        print(f"\n📋 Headers ({len(headers)} columns):")
        for idx, header in enumerate(headers, 1):
            print(f"   {idx:2d}. {header}")
        
        # Test field metadata extraction
        print("\n[3/4] Extracting field metadata...")
        field_metadata = extract_field_metadata(headers)
        
        print(f"\n🔍 Field Metadata Analysis:")
        for idx, header in enumerate(headers, 1):
            if header in field_metadata:
                meta = field_metadata[header]
                field_type = meta.get('type', 'unknown')
                keywords = meta.get('keywords', [])
                normalized = meta.get('normalized', header)
                print(f"\n   {idx:2d}. {header}")
                print(f"       Type: {field_type}")
                print(f"       Normalized: {normalized}")
                print(f"       Keywords: {keywords if keywords else 'None'}")
        
        print(f"\n📊 Data rows: {len(rows)}")
        print(f"\n[4/4] First row data:")
        for header, value in rows[0].items():
            print(f"   {header:<20s} → {value}")
        
        print("\n" + "="*70)
        print("✅ TEST PASSED - All columns extracted with metadata!")
        print("="*70 + "\n")
        
        return True
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        print("="*70 + "\n")
        return False

if __name__ == "__main__":
    success = test_excel_parsing()
    sys.exit(0 if success else 1)

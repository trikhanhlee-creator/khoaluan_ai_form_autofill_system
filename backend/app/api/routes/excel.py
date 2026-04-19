"""
API routes for Excel file handling
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query, Body, Form
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, inspect
import openpyxl
from io import BytesIO
import csv
import os
import zipfile
import xlrd
import re
import json
import unicodedata
from datetime import datetime
from typing import Any
from uuid import uuid4
from app.core.logger import logger
from app.core.auth import get_current_user
from app.db.session import get_db
from app.db.models import User, Entry, ExcelTemplate

router = APIRouter(prefix="/api/excel", tags=["excel"])

# Store uploaded Excel data in memory (in production, use a database)
excel_data_store = {}


def _normalize_lookup_key(value: str) -> str:
    return _normalize_match_key(value)


def _normalize_match_key(value: str) -> str:
    """Normalize text for fuzzy header-field matching (supports Vietnamese accents)."""
    text = (value or "").strip().lower()
    if not text:
        return ""

    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _match_header_score(field_candidate: str, header: str) -> int:
    field_key = _normalize_match_key(field_candidate)
    header_key = _normalize_match_key(header)

    if not field_key or not header_key:
        return 0

    if field_key == header_key:
        return 100

    if field_key in header_key or header_key in field_key:
        overlap_penalty = abs(len(field_key) - len(header_key))
        return max(70, 92 - overlap_penalty)

    field_tokens = set(field_key.split())
    header_tokens = set(header_key.split())
    if not field_tokens or not header_tokens:
        return 0

    overlap = field_tokens.intersection(header_tokens)
    if not overlap:
        return 0

    base = int((len(overlap) / max(len(field_tokens), len(header_tokens))) * 80)
    if field_tokens.issubset(header_tokens) or header_tokens.issubset(field_tokens):
        base += 10
    return min(base, 95)


def _extract_row_values_from_header(rows: list[dict], header: str) -> dict[str, str]:
    row_values: dict[str, str] = {}
    for row_index, row in enumerate(rows):
        if not isinstance(row, dict):
            continue

        value = str(row.get(header, "") or "").strip()
        if not value:
            continue

        row_values[str(row_index)] = value

    return row_values


def _build_distinct_choices_from_row_values(row_values: dict[str, str], limit: int = 200) -> list[dict[str, Any]]:
    seen: set[str] = set()
    choices: list[dict[str, Any]] = []

    def _sort_key(item: str) -> tuple[int, str]:
        try:
            return (0, f"{int(item):09d}")
        except Exception:
            return (1, item)

    for row_key in sorted(row_values.keys(), key=_sort_key):
        value = str(row_values.get(row_key, "") or "").strip()
        if not value or value in seen:
            continue

        seen.add(value)
        try:
            row_index = int(row_key)
        except Exception:
            row_index = row_key

        choices.append({
            "value": value,
            "row_index": row_index,
        })

        if len(choices) >= limit:
            break

    return choices


def _is_unique_row_values(row_values: dict[str, str]) -> bool:
    values = [str(v).strip() for v in row_values.values() if str(v).strip()]
    if not values:
        return False
    return len(values) == len(set(values))


def _normalize_reference_targets(raw_targets: Any) -> list[dict[str, Any]]:
    targets: list[dict[str, Any]] = []
    if not isinstance(raw_targets, list):
        return targets

    for idx, item in enumerate(raw_targets):
        key = ""
        display = ""
        candidates: list[str] = []

        if isinstance(item, str):
            text = item.strip()
            if not text:
                continue
            key = text
            display = text
            candidates = [text]
        elif isinstance(item, dict):
            key = str(
                item.get("field_id")
                or item.get("key")
                or item.get("field_name")
                or item.get("name")
                or item.get("field_label")
                or item.get("label")
                or f"field_{idx}"
            ).strip()

            display = str(
                item.get("field_label")
                or item.get("label")
                or item.get("field_name")
                or item.get("name")
                or key
            ).strip()

            candidates = [
                str(item.get("field_label") or "").strip(),
                str(item.get("label") or "").strip(),
                str(item.get("field_name") or "").strip(),
                str(item.get("name") or "").strip(),
                display,
            ]
            candidates = [c for c in candidates if c]

            if not key:
                key = display or f"field_{idx}"
        else:
            continue

        dedup_candidates: list[str] = []
        seen_candidates: set[str] = set()
        for candidate in candidates:
            normalized = _normalize_match_key(candidate)
            if not normalized or normalized in seen_candidates:
                continue
            seen_candidates.add(normalized)
            dedup_candidates.append(candidate)

        if not dedup_candidates:
            continue

        targets.append({
            "key": key,
            "display": display or key,
            "candidates": dedup_candidates,
        })

    return targets


def _is_family_name_header(normalized: str) -> bool:
    key = f" {normalized} "
    tokens = set(normalized.split())
    if "ten" in tokens and "ho" not in tokens:
        return False

    if key.strip() in {"ho", "ho dem", "ho lot", "surname", "last name", "family name"}:
        return True

    return "ho" in tokens and "ten" not in tokens


def _is_given_name_header(normalized: str) -> bool:
    key = f" {normalized} "
    tokens = set(normalized.split())
    if "ho" in tokens and "ten" not in tokens:
        return False

    if key.strip() in {"ten", "given name", "first name", "name"}:
        return True

    return "ten" in tokens and "ho" not in tokens


def _is_middle_name_header(normalized: str) -> bool:
    key = f" {normalized} "
    tokens = set(normalized.split())
    if key.strip() in {"dem", "ten dem", "middle", "middle name"}:
        return True
    return "dem" in tokens or "lot" in tokens


def _is_full_name_header(normalized: str) -> bool:
    key = f" {normalized} "
    compact = key.strip()
    if compact in {"ho va ten", "ho ten", "ten day du", "full name", "fullname"}:
        return True

    tokens = set(compact.split())
    if {"ho", "ten"}.issubset(tokens):
        return True

    return False


def _is_full_name_target(target: dict[str, Any]) -> bool:
    candidates = [target.get("display", "")]
    candidates.extend(target.get("candidates", []) if isinstance(target.get("candidates"), list) else [])

    for text in candidates:
        normalized = _normalize_match_key(str(text))
        if not normalized:
            continue

        if "ho va ten" in normalized or "ho ten" in normalized:
            return True
        if "full name" in normalized or "ten day du" in normalized:
            return True
        if "ho" in normalized.split() and "ten" in normalized.split():
            return True

    return False


def _find_explicit_full_name_header(headers: list[str]) -> str | None:
    for header in headers:
        normalized = _normalize_lookup_key(header)
        if normalized and _is_full_name_header(normalized):
            return header
    return None


def _find_name_part_headers(headers: list[str]) -> tuple[str | None, str | None, str | None]:
    family_header: str | None = None
    middle_header: str | None = None
    given_header: str | None = None

    for header in headers:
        normalized = _normalize_lookup_key(header)
        if not normalized:
            continue

        if family_header is None and _is_family_name_header(normalized):
            family_header = header
            continue

        if middle_header is None and _is_middle_name_header(normalized):
            middle_header = header
            continue

        if given_header is None and _is_given_name_header(normalized):
            given_header = header

    return family_header, middle_header, given_header


def _extract_composite_full_name_row_values(
    rows: list[dict],
    family_header: str,
    given_header: str,
    middle_header: str | None = None,
    limit: int = 2000,
) -> dict[str, str]:
    row_values: dict[str, str] = {}

    for row_index, row in enumerate(rows):
        if not isinstance(row, dict):
            continue

        family_name = str(row.get(family_header, "") or "").strip()
        middle_name = str(row.get(middle_header, "") or "").strip() if middle_header else ""
        given_name = str(row.get(given_header, "") or "").strip()

        full_name = " ".join(part for part in [family_name, middle_name, given_name] if part).strip()
        if not full_name:
            continue

        row_values[str(row_index)] = full_name
        if len(row_values) >= limit:
            break

    return row_values


def _get_or_create_excel_form_id(db: Session, user_id: int) -> int:
    bind = db.get_bind()
    inspector = inspect(bind)
    if not inspector.has_table("forms"):
        raise HTTPException(status_code=500, detail="Thiếu bảng forms")

    columns = {col["name"] for col in inspector.get_columns("forms")}

    select_sql = "SELECT id FROM forms WHERE user_id = :user_id"
    if "form_type" in columns:
        select_sql += " AND form_type = 'excel'"
    select_sql += " ORDER BY id ASC LIMIT 1"

    row = db.execute(text(select_sql), {"user_id": user_id}).first()
    if row:
        return int(row[0])

    insert_columns = ["user_id", "form_name"]
    insert_values = [":user_id", ":form_name"]
    params = {
        "user_id": user_id,
        "form_name": "Excel Smart Form",
    }

    if "description" in columns:
        insert_columns.append("description")
        insert_values.append(":description")
        params["description"] = "System form for Excel field mapping"
    if "form_type" in columns:
        insert_columns.append("form_type")
        insert_values.append(":form_type")
        params["form_type"] = "excel"
    if "is_template" in columns:
        insert_columns.append("is_template")
        insert_values.append(":is_template")
        params["is_template"] = 1
    if "created_at" in columns:
        insert_columns.append("created_at")
        insert_values.append("CURRENT_TIMESTAMP")
    if "updated_at" in columns:
        insert_columns.append("updated_at")
        insert_values.append("CURRENT_TIMESTAMP")

    insert_sql = f"INSERT INTO forms ({', '.join(insert_columns)}) VALUES ({', '.join(insert_values)})"
    result = db.execute(text(insert_sql), params)
    db.commit()
    if result.lastrowid:
        return int(result.lastrowid)

    row = db.execute(text(select_sql), {"user_id": user_id}).first()
    if not row:
        raise HTTPException(status_code=500, detail="Không thể tạo form Excel")
    return int(row[0])


def _ensure_excel_fields(db: Session, form_id: int, headers: list[str]) -> dict[str, int]:
    bind = db.get_bind()
    inspector = inspect(bind)
    if not inspector.has_table("fields"):
        raise HTTPException(status_code=500, detail="Thiếu bảng fields")

    columns = {col["name"] for col in inspector.get_columns("fields")}
    rows = db.execute(
        text("SELECT id, field_name FROM fields WHERE form_id = :form_id"),
        {"form_id": form_id}
    ).all()
    field_map = {str(row[1]): int(row[0]) for row in rows if row[1]}

    for idx, header in enumerate(headers):
        name = (header or "").strip()
        if not name or name in field_map:
            continue

        insert_columns = ["form_id", "field_name"]
        insert_values = [":form_id", ":field_name"]
        params = {"form_id": form_id, "field_name": name}

        if "field_type" in columns:
            insert_columns.append("field_type")
            insert_values.append(":field_type")
            params["field_type"] = "text"
        if "display_order" in columns:
            insert_columns.append("display_order")
            insert_values.append(":display_order")
            params["display_order"] = idx
        if "is_required" in columns:
            insert_columns.append("is_required")
            insert_values.append(":is_required")
            params["is_required"] = 0
        if "created_at" in columns:
            insert_columns.append("created_at")
            insert_values.append("CURRENT_TIMESTAMP")

        result = db.execute(
            text(f"INSERT INTO fields ({', '.join(insert_columns)}) VALUES ({', '.join(insert_values)})"),
            params,
        )
        if result.lastrowid:
            field_map[name] = int(result.lastrowid)

    db.commit()

    if any((header or "").strip() and (header or "").strip() not in field_map for header in headers):
        rows = db.execute(
            text("SELECT id, field_name FROM fields WHERE form_id = :form_id"),
            {"form_id": form_id}
        ).all()
        field_map.update({str(row[1]): int(row[0]) for row in rows if row[1]})

    return field_map


def _can_access_session(current_user: User, session_owner_id: int) -> bool:
    """Owner can access their own session; admin can access all."""
    return current_user.is_admin or current_user.id == session_owner_id


SESSION_TEMPLATE_PREFIX = "excel_session::"


def _session_template_name(filename: str) -> str:
    return f"{SESSION_TEMPLATE_PREFIX}{filename}"


def _normalize_history_items(history: Any) -> list[dict]:
    if not isinstance(history, list):
        return []
    normalized = []
    for item in history:
        if not isinstance(item, dict):
            continue
        history_id = str(item.get("history_id") or "").strip()
        row_index = item.get("row_index")
        row_data = item.get("row_data")
        saved_at = str(item.get("saved_at") or "").strip()
        if not history_id:
            history_id = uuid4().hex
        try:
            row_index = int(row_index)
        except Exception:
            row_index = -1
        if row_index < 0:
            continue
        if not isinstance(row_data, dict):
            row_data = {}
        if not saved_at:
            saved_at = datetime.utcnow().isoformat()
        normalized.append({
            "history_id": history_id,
            "row_index": row_index,
            "row_data": row_data,
            "saved_at": saved_at,
        })
    return normalized


def _append_session_history(data: dict, row_index: int, row_data: dict) -> dict:
    history = _normalize_history_items(data.get("history"))
    item = {
        "history_id": uuid4().hex,
        "row_index": int(row_index),
        "row_data": row_data,
        "saved_at": datetime.utcnow().isoformat(),
    }
    history.append(item)
    data["history"] = history
    return item


def _persist_session_to_db(
    db: Session,
    session_id: str,
    user_id: int,
    filename: str,
    headers: list,
    rows: list,
    created_at: str,
    history: list | None = None,
) -> None:
    normalized_history = _normalize_history_items(history)
    persisted_payload = {
        "rows": rows,
        "field_metadata": extract_field_metadata(headers),
        "created_at": created_at,
        "history": normalized_history,
    }

    record = db.query(ExcelTemplate).filter(
        ExcelTemplate.user_id == user_id,
        ExcelTemplate.file_path == session_id,
        ExcelTemplate.template_name.like(f"{SESSION_TEMPLATE_PREFIX}%")
    ).first()

    if record is None:
        record = ExcelTemplate(
            user_id=user_id,
            template_name=_session_template_name(filename),
            file_path=session_id,
            original_filename=filename,
            sheet_name="Sheet1",
            headers_json=json.dumps(headers, ensure_ascii=False),
            data_row_start=len(rows),
            mapping_json=json.dumps(persisted_payload, ensure_ascii=False),
        )
        db.add(record)
    else:
        record.template_name = _session_template_name(filename)
        record.original_filename = filename
        record.headers_json = json.dumps(headers, ensure_ascii=False)
        record.data_row_start = len(rows)
        record.mapping_json = json.dumps(persisted_payload, ensure_ascii=False)

    db.commit()


def _load_session_from_db(db: Session, session_id: str) -> dict | None:
    record = db.query(ExcelTemplate).filter(
        ExcelTemplate.file_path == session_id,
        ExcelTemplate.template_name.like(f"{SESSION_TEMPLATE_PREFIX}%")
    ).first()

    if not record:
        return None

    try:
        headers = json.loads(record.headers_json) if record.headers_json else []
    except Exception:
        headers = []

    persisted_payload = {}
    try:
        persisted_payload = json.loads(record.mapping_json) if record.mapping_json else {}
    except Exception:
        persisted_payload = {}

    rows = persisted_payload.get("rows", []) if isinstance(persisted_payload, dict) else []
    field_metadata = persisted_payload.get("field_metadata") if isinstance(persisted_payload, dict) else None
    created_at = persisted_payload.get("created_at") if isinstance(persisted_payload, dict) else None
    history = persisted_payload.get("history") if isinstance(persisted_payload, dict) else []

    if not isinstance(rows, list):
        rows = []
    if not isinstance(field_metadata, dict):
        field_metadata = extract_field_metadata(headers)
    if not created_at:
        created_at = record.created_at.isoformat() if record.created_at else datetime.utcnow().isoformat()
    history = _normalize_history_items(history)

    return {
        "user_id": int(record.user_id),
        "filename": record.original_filename or record.template_name.replace(SESSION_TEMPLATE_PREFIX, ""),
        "headers": headers,
        "rows": rows,
        "total_rows": len(rows),
        "created_at": str(created_at),
        "field_metadata": field_metadata,
        "history": history,
    }


def _update_session_rows_in_db(db: Session, session_id: str, rows: list) -> None:
    record = db.query(ExcelTemplate).filter(
        ExcelTemplate.file_path == session_id,
        ExcelTemplate.template_name.like(f"{SESSION_TEMPLATE_PREFIX}%")
    ).first()
    if not record:
        return

    payload = {}
    try:
        payload = json.loads(record.mapping_json) if record.mapping_json else {}
    except Exception:
        payload = {}

    if not isinstance(payload, dict):
        payload = {}

    payload["rows"] = rows
    if "created_at" not in payload:
        payload["created_at"] = record.created_at.isoformat() if record.created_at else datetime.utcnow().isoformat()
    if "field_metadata" not in payload:
        try:
            payload["field_metadata"] = extract_field_metadata(json.loads(record.headers_json) if record.headers_json else [])
        except Exception:
            payload["field_metadata"] = {}
    if "history" not in payload:
        payload["history"] = []

    record.mapping_json = json.dumps(payload, ensure_ascii=False)
    record.data_row_start = len(rows)
    db.commit()


def _get_session_data_or_404(session_id: str, current_user: User, db: Session | None = None) -> dict:
    if session_id not in excel_data_store and db is not None:
        persisted = _load_session_from_db(db, session_id)
        if persisted is not None:
            excel_data_store[session_id] = persisted

    if session_id not in excel_data_store:
        raise HTTPException(status_code=404, detail="Session not found")

    data = excel_data_store[session_id]
    owner_id = data.get('user_id')
    if owner_id is None or not _can_access_session(current_user, owner_id):
        raise HTTPException(status_code=403, detail="Không có quyền truy cập session này")

    return data

def detect_field_type(field_name: str) -> str:
    """Detect field type from field name"""
    field_lower = field_name.lower()
    
    # Pattern detection
    if re.search(r'(%|phần trăm|percentage)', field_lower):
        return 'percentage'
    elif re.search(r'(ngày|date|tháng|năm)', field_lower):
        return 'date'
    elif re.search(r'(điểm|gpa|score|mark)', field_lower):
        return 'number'
    elif re.search(r'(mã|code|id)', field_lower):
        return 'code'
    elif re.search(r'(email|address|địa chỉ)', field_lower):
        return 'text_long'
    else:
        return 'text'

def normalize_field_name(field_name: str) -> str:
    """Normalize field name by removing extra spaces and formatting"""
    # Remove multiple spaces
    field_name = re.sub(r'\s+', ' ', field_name).strip()
    # Remove trailing numbers or special chars from percentage labels
    field_name = re.sub(r'\s*\(\d+%\)\s*$', '', field_name).strip()
    # Fix common encoding issues and typos
    # Ho đêm -> Họ (family name)
    field_name = re.sub(r'Ho\s+đêm', 'Họ', field_name, flags=re.IGNORECASE)
    # Họ đêm -> Họ (if it's a typo)
    field_name = re.sub(r'Họ\s+đêm', 'Họ', field_name, flags=re.IGNORECASE)
    # Remove trailing numbers that often appear in data
    field_name = re.sub(r'^\d+[\s\.]', '', field_name).strip()
    return field_name

def is_label_row(row_values: list) -> bool:
    """Check if a row contains only labels/formatting (e.g., percentage signs) with mostly empty cells
    Returns True if row should be skipped
    """
    if not row_values:
        return False
    
    # Count meaningful data (not just empty, not just percentage markers)
    meaningful_cells = 0
    for val in row_values:
        if val is None or val == '':
            continue
        val_str = str(val).strip()
        # Skip if it's just a percentage marker or unit label
        if val_str and not re.match(r'^[\d\s%()%]+$', val_str):
            meaningful_cells += 1
    
    # If less than 20% of cells have meaningful data, it's likely a label row
    non_empty_cells = sum(1 for v in row_values if v is not None and str(v).strip())
    if non_empty_cells == 0:
        return False
    
    return meaningful_cells < non_empty_cells * 0.2

def looks_like_field_header_row(row_values: list) -> bool:
    """Check if a row looks like actual field headers (not section titles)
    Looks for Vietnamese field name patterns like: STT, Mã, Tên, Họ, Ngày, Lớp, etc.
    """
    if not row_values or len(row_values) < 3:
        return False
    
    # Common Vietnamese field name keywords
    field_keywords = [
        'stt', 'mã', 'tên', 'họ', 'ngày', 'lớp', 'giới tính', 'tuổi', 'địa chỉ',
        'điểm', 'gpa', 'kết quả', 'chuyên cần', 'thường xuyên', 'quân sự',
        'tình nguyện', 'học tập', 'hạnh kiểm', 'email', 'điện thoại',
        'sinh viên', 'khoa', 'năm', 'tháng', 'chính'
    ]
    
    matched_keywords = 0
    total_text_cells = 0
    
    for cell_value in row_values:
        if cell_value is None or cell_value == '':
            continue
        
        cell_str = str(cell_value).strip().lower()
        
        # Count text cells
        if any(c.isalpha() for c in cell_str):
            total_text_cells += 1
            
            # Check if this cell contains any field keyword
            for keyword in field_keywords:
                if keyword in cell_str:
                    matched_keywords += 1
                    break
    
    # If we found at least 30% of cells matching field keywords, likely a real header
    if total_text_cells > 0:
        return matched_keywords / total_text_cells >= 0.3
    
    return False

def looks_like_header_row(row_values: list) -> bool:
    """Check if a row looks like a header row (mostly text, no pure numbers)"""
    if not row_values:
        return False
    
    text_cells = 0
    pure_number_cells = 0
    
    for cell_value in row_values:
        if cell_value is None or cell_value == '':
            continue
        
        cell_str = str(cell_value).strip()
        
        # Check if cell contains mostly text (contains letters)
        if any(c.isalpha() for c in cell_str):
            text_cells += 1
        # Check if cell is purely numeric (suggests data row)
        elif re.match(r'^\d+(\.\d+)?$', cell_str):
            pure_number_cells += 1
    
    # Header rows should have more text cells than pure numbers
    total_non_empty = text_cells + pure_number_cells
    if total_non_empty == 0:
        return False
    
    # If >60% are text cells, likely a header row
    return text_cells / total_non_empty > 0.6

def extract_field_metadata(headers: list) -> dict:
    """Extract metadata about fields for better grouping"""
    metadata = {}
    
    for header in headers:
        normalized = normalize_field_name(header)
        metadata[header] = {
            'original': header,
            'normalized': normalized,
            'type': detect_field_type(header),
            'keywords': extract_keywords(header)
        }
    
    return metadata

def extract_keywords(field_name: str) -> list:
    """Extract keywords from field name"""
    keywords = []
    
    # Vietnamese keywords mapping - support both common variations
    keyword_patterns = {
        'thông tin': ['mã', 'họ', 'tên', 'giới tính', 'ngày sinh', 'lớp', 'khoa', 'sinh viên'],
        'chuyên cần': ['chuyên cần', 'dự', 'vắng', '20%', 'attendance'],
        'thường xuyên': ['thường xuyên', 'thương xuyên', '30%', 'hành động', 'dẫn đầu'],
        'quân sự': ['quân sự', 'military', 'huấn luyện', 'quân'],
        'tình nguyện': ['tình nguyện', 'volunteer', 'xã hội', 'công tác'],
        'học tập': ['học tập', 'điểm', 'gpa', 'kết quả', 'tổng kết'],
        'hạnh kiểm': ['hạnh kiểm', 'conduct', 'kỷ luật', 'rút kinh nghiệm']
    }
    
    field_lower = field_name.lower()
    for category, keywords_list in keyword_patterns.items():
        for keyword in keywords_list:
            if keyword in field_lower:
                keywords.append(category)
                break
    
    return keywords

def is_valid_xlsx(file_content: bytes) -> tuple[bool, str]:
    """Check if file is a valid XLSX file"""
    try:
        # XLSX files are ZIP archives, check magic number
        if len(file_content) < 4:
            return False, "File too small"
        
        # Check ZIP magic number (PK..)
        if file_content[:2] != b'PK':
            return False, "Not a valid XLSX file (not a ZIP archive)"
        
        # Try to open as ZIP
        try:
            with zipfile.ZipFile(BytesIO(file_content)) as zf:
                # Check for required XLSX files
                if '[Content_Types].xml' not in zf.namelist():
                    return False, "Missing [Content_Types].xml in XLSX archive"
                
                # Check for workbook
                if 'xl/workbook.xml' not in zf.namelist():
                    return False, "Missing xl/workbook.xml - possible corrupted XLSX"
                
                return True, "Valid XLSX"
        except zipfile.BadZipFile:
            return False, "Corrupted ZIP/XLSX file"
    except Exception as e:
        return False, f"ZIP validation error: {str(e)}"

def parse_excel_with_openpyxl(excel_file: BytesIO) -> tuple[list, list, str]:
    """Parse Excel file with openpyxl - extract ALL columns"""
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
        worksheet = workbook.active
        
        if worksheet is None:
            raise ValueError("No active worksheet found")
        
        # Find header row with improved detection
        header_row_idx = 1
        
        # First priority: Find row with MOST actual field keywords (for multi-row headers)
        # Count field keyword matches for each row
        best_keyword_count = 0
        best_non_empty_count = 0
        
        for row_num in range(1, min(worksheet.max_row + 1, 10)):
            row_values = []
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row_num, col_idx).value
                row_values.append(cell_value)
            
            # Count field keywords in this row
            field_keywords = [
                'stt', 'mã', 'tên', 'họ', 'đệm', 'ngày', 'lớp', 'giới tính', 'tuổi', 'địa chỉ',
                'điểm', 'gpa', 'kết quả', 'chuyên cần', 'thường xuyên', 'thương xuyên', 'quân sự',
                'tình nguyện', 'học tập', 'hạnh kiểm', 'email', 'điện thoại',
                'sinh viên', 'khoa', 'năm', 'tháng', 'chính'
            ]
            
            keyword_count = 0
            non_empty_count = 0
            for cell_value in row_values:
                if cell_value is not None and str(cell_value).strip():
                    non_empty_count += 1
                    cell_str = str(cell_value).strip().lower()
                    for keyword in field_keywords:
                        if keyword in cell_str:
                            keyword_count += 1
                            break
            
            # Prefer row with more field keywords, then more non-empty cells
            if keyword_count > best_keyword_count or (keyword_count == best_keyword_count and non_empty_count > best_non_empty_count):
                best_keyword_count = keyword_count
                best_non_empty_count = non_empty_count
                header_row_idx = row_num
        
        # Fallback: if no keywords found, use row with most non-empty cells
        if best_keyword_count == 0:
            header_row_idx = 1
            max_cells = 0
            
            for row_num in range(1, min(worksheet.max_row + 1, 10)):
                row_values = []
                cell_count = 0
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row_num, col_idx).value
                    row_values.append(cell_value)
                    if cell_value is not None and str(cell_value).strip():
                        cell_count += 1
                
                if looks_like_header_row(row_values) and cell_count > max_cells:
                    max_cells = cell_count
                    header_row_idx = row_num
        
        # Extract headers - get ALL columns
        headers = []
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(header_row_idx, col_idx).value
            
            # If current row has empty cell, try to get from row above (for merged header cases)
            merged_value = None
            if (cell_value is None or cell_value == '') and header_row_idx > 1:
                # Try to get value from previous rows
                for prev_row in range(header_row_idx - 1, 0, -1):
                    prev_value = worksheet.cell(prev_row, col_idx).value
                    if prev_value is not None and str(prev_value).strip():
                        merged_value = prev_value
                        break
            
            final_value = cell_value if (cell_value is not None and str(cell_value).strip()) else merged_value
            
            if final_value is not None:
                # Normalize the header name
                header = normalize_field_name(str(final_value).strip())
                if header:  # Only add if not empty after normalization
                    headers.append(header)
                else:
                    headers.append(f"Column_{col_idx}")
            else:
                headers.append(f"Column_{col_idx}")
        
        if not headers or all(h.startswith("Column_") for h in headers):
            raise ValueError("Không tìm thấy header trong file. Vui lòng đảm bảo hàng đầu tiên có tiêu đề cột.")
        
        # Get data rows (starting after header row)
        # Skip any label-only rows immediately after headers
        start_row = header_row_idx + 1
        
        # Check if first row after header is a label row (like percentage markers)
        if start_row <= worksheet.max_row:
            first_data_row_values = []
            for col_idx in range(1, worksheet.max_column + 1):
                first_data_row_values.append(worksheet.cell(start_row, col_idx).value)
            
            if is_label_row(first_data_row_values):
                start_row += 1  # Skip this label row
        
        rows = []
        for row_num in range(start_row, worksheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row_num, col_idx)
                value = cell.value
                
                if value is not None and value != '':
                    has_data = True
                    # Handle different types properly
                    if isinstance(value, (int, float)):
                        if isinstance(value, float) and value.is_integer():
                            row_data[header] = str(int(value))
                        else:
                            row_data[header] = str(value)
                    else:
                        row_data[header] = str(value)
                else:
                    row_data[header] = ""
            
            # Add row if it has any data
            if has_data:
                rows.append(row_data)
        
        if not rows:
            raise ValueError("Không tìm thấy dữ liệu trong file. Vui lòng kiểm tra file Excel.")
        
        msg = f"Tải file thành công! Tìm thấy {len(rows)} dòng dữ liệu"
        return headers, rows, msg
    except Exception as e:
        error_msg = str(e)
        # Provide helpful Vietnamese messages
        if "Không tìm thấy" in error_msg:
            raise Exception(f"Lỗi: {error_msg}")
        else:
            raise Exception(f"Lỗi: Không thể đọc file Excel: {error_msg}")

def parse_xls_file(file_content: bytes) -> tuple[list, list, str]:
    """Parse .xls file using xlrd - extract ALL columns"""
    try:
        # Open XLS file
        workbook = xlrd.open_workbook(file_contents=file_content, on_demand=True)
        
        if not workbook.nsheets:
            raise ValueError("No sheets found in XLS file")
        
        # Get first sheet
        worksheet = workbook.sheet_by_index(0)
        
        if worksheet.nrows == 0:
            raise ValueError("Worksheet is empty")
        
        # Find header row with improved detection
        header_row_idx = 0
        
        # First priority: Find row with MOST actual field keywords (for multi-row headers)
        # Count field keyword matches for each row
        best_keyword_count = 0
        best_non_empty_count = 0
        
        for row_num in range(min(worksheet.nrows, 10)):
            row_values = []
            for col_idx in range(worksheet.ncols):
                cell_value = worksheet.cell_value(row_num, col_idx)
                row_values.append(cell_value)
            
            # Count field keywords in this row
            field_keywords = [
                'stt', 'mã', 'tên', 'họ', 'đệm', 'ngày', 'lớp', 'giới tính', 'tuổi', 'địa chỉ',
                'điểm', 'gpa', 'kết quả', 'chuyên cần', 'thường xuyên', 'thương xuyên', 'quân sự',
                'tình nguyện', 'học tập', 'hạnh kiểm', 'email', 'điện thoại',
                'sinh viên', 'khoa', 'năm', 'tháng', 'chính'
            ]
            
            keyword_count = 0
            non_empty_count = 0
            for cell_value in row_values:
                if cell_value is not None and str(cell_value).strip():
                    non_empty_count += 1
                    cell_str = str(cell_value).strip().lower()
                    for keyword in field_keywords:
                        if keyword in cell_str:
                            keyword_count += 1
                            break
            
            # Prefer row with more field keywords, then more non-empty cells
            if keyword_count > best_keyword_count or (keyword_count == best_keyword_count and non_empty_count > best_non_empty_count):
                best_keyword_count = keyword_count
                best_non_empty_count = non_empty_count
                header_row_idx = row_num
        
        # Fallback: if no keywords found, use row with most non-empty cells
        if best_keyword_count == 0:
            header_row_idx = 0
            max_cells = 0
            
            for row_num in range(min(worksheet.nrows, 10)):
                row_values = []
                cell_count = 0
                for col_idx in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row_num, col_idx)
                    row_values.append(cell_value)
                    if cell_value is not None and str(cell_value).strip():
                        cell_count += 1
                
                if looks_like_header_row(row_values) and cell_count > max_cells:
                    max_cells = cell_count
                    header_row_idx = row_num
        
        # Extract headers - get ALL columns
        headers = []
        for col_idx in range(worksheet.ncols):
            cell_value = worksheet.cell_value(header_row_idx, col_idx)
            
            # If current row has empty cell, try to get from row above (for merged header cases)
            merged_value = None
            if (cell_value is None or cell_value == '') and header_row_idx > 0:
                # Try to get value from previous rows
                for prev_row in range(header_row_idx - 1, -1, -1):
                    prev_value = worksheet.cell_value(prev_row, col_idx)
                    if prev_value is not None and str(prev_value).strip():
                        merged_value = prev_value
                        break
            
            final_value = cell_value if (cell_value is not None and str(cell_value).strip()) else merged_value
            
            if final_value is not None:
                # Normalize the header name
                header = normalize_field_name(str(final_value).strip())
                if header:  # Only add if not empty after normalization
                    headers.append(header)
                else:
                    headers.append(f"Column_{col_idx + 1}")
            else:
                headers.append(f"Column_{col_idx + 1}")
        
        if not headers or all(h.startswith("Column_") for h in headers):
            raise ValueError("Không tìm thấy header trong file. Vui lòng đảm bảo hàng đầu tiên có tiêu đề cột.")
        
        # Get data rows (starting after header row)
        # Skip any label-only rows immediately after headers
        start_row = header_row_idx + 1
        
        # Check if first row after header is a label row (like percentage markers)
        if start_row < worksheet.nrows:
            first_data_row_values = []
            for col_idx in range(worksheet.ncols):
                first_data_row_values.append(worksheet.cell_value(start_row, col_idx))
            
            if is_label_row(first_data_row_values):
                start_row += 1  # Skip this label row
        
        rows = []
        for row_idx in range(start_row, worksheet.nrows):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers):
                cell_value = worksheet.cell_value(row_idx, col_idx)
                # Handle different cell types and numbers
                if cell_value is not None and cell_value != '':
                    has_data = True
                    # Convert to string, handling numbers properly
                    if isinstance(cell_value, (int, float)):
                        # Check if it's a float that should be an int
                        if isinstance(cell_value, float) and cell_value.is_integer():
                            row_data[header] = str(int(cell_value))
                        else:
                            row_data[header] = str(cell_value)
                    else:
                        row_data[header] = str(cell_value)
                else:
                    row_data[header] = ""
            
            # Add row if it has any data
            if has_data:
                rows.append(row_data)
        
        if not rows:
            raise ValueError("Không tìm thấy dữ liệu trong file. Vui lòng kiểm tra file Excel.")
        
        msg = f"Tải file thành công! Tìm thấy {len(rows)} dòng dữ liệu"
        return headers, rows, msg
    except Exception as e:
        error_msg = str(e)
        if "Không tìm thấy" in error_msg:
            raise Exception(f"Lỗi: {error_msg}")
        else:
            raise Exception(f"Lỗi: Không thể đọc file XLS: {error_msg}")


@router.post("/reference-field-options")
async def get_reference_field_options(
    file: UploadFile = File(...),
    fields_json: str = Form("[]"),
    current_user: User = Depends(get_current_user)
):
    """Read a reference Excel file and return matched dropdown options by field."""
    _ = current_user

    try:
        filename = file.filename or ""
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext not in ('.xlsx', '.xls'):
            raise HTTPException(
                status_code=400,
                detail="Chi hỗ trợ file Excel tham chiếu (.xlsx, .xls)"
            )

        raw_content = await file.read()
        if not raw_content:
            raise HTTPException(status_code=400, detail="File tham chiếu rỗng hoặc không hợp lệ")

        if file_ext == '.xlsx':
            is_valid, error_msg = is_valid_xlsx(raw_content)
            if not is_valid:
                raise HTTPException(status_code=400, detail=f"File Excel không hợp lệ: {error_msg}")

        try:
            parsed_targets = json.loads(fields_json or "[]")
        except Exception:
            raise HTTPException(status_code=400, detail="fields_json không hợp lệ")

        targets = _normalize_reference_targets(parsed_targets)
        if not targets:
            return JSONResponse({
                "status": "success",
                "filename": filename,
                "matches": {},
                "unmatched_keys": [],
                "matched_count": 0,
                "message": "Không có trường mục tiêu để đối chiếu"
            })

        if file_ext == '.xlsx':
            headers, rows, _ = parse_excel_with_openpyxl(BytesIO(raw_content))
        else:
            headers, rows, _ = parse_xls_file(raw_content)

        matches: dict[str, dict[str, Any]] = {}
        unmatched_keys: list[str] = []
        explicit_full_name_header = _find_explicit_full_name_header(headers)
        family_header, middle_header, given_header = _find_name_part_headers(headers)

        for target in targets:
            if _is_full_name_target(target):
                full_name_row_values: dict[str, str] = {}
                full_name_header_label = ""
                matched_by = "full_name_composition"

                if explicit_full_name_header:
                    full_name_row_values = _extract_row_values_from_header(rows, explicit_full_name_header)
                    if full_name_row_values:
                        full_name_header_label = explicit_full_name_header
                        matched_by = explicit_full_name_header

                if (not full_name_row_values) and family_header and given_header:
                    full_name_row_values = _extract_composite_full_name_row_values(
                        rows=rows,
                        family_header=family_header,
                        middle_header=middle_header,
                        given_header=given_header,
                    )
                    if full_name_row_values:
                        full_name_header_label = (
                            f"{family_header} + {middle_header} + {given_header}"
                            if middle_header
                            else f"{family_header} + {given_header}"
                        )
                        matched_by = "full_name_composition"

                if full_name_row_values:
                    full_name_choices = _build_distinct_choices_from_row_values(full_name_row_values)
                    full_name_values = [item["value"] for item in full_name_choices]
                    matches[target["key"]] = {
                        "field": target["display"],
                        "matched_header": full_name_header_label,
                        "matched_by": matched_by,
                        "score": 97,
                        "values": full_name_values,
                        "choices": full_name_choices,
                        "row_values": full_name_row_values,
                        "is_unique": _is_unique_row_values(full_name_row_values),
                        "total_values": len(full_name_values),
                    }
                    continue

            best_header = ""
            best_score = 0
            best_candidate = ""

            for candidate in target["candidates"]:
                for header in headers:
                    score = _match_header_score(candidate, header)
                    if score > best_score:
                        best_score = score
                        best_header = header
                        best_candidate = candidate

            if not best_header or best_score < 55:
                unmatched_keys.append(target["key"])
                continue

            row_values = _extract_row_values_from_header(rows, best_header)
            choices = _build_distinct_choices_from_row_values(row_values)
            values = [item["value"] for item in choices]
            if not values:
                unmatched_keys.append(target["key"])
                continue

            matches[target["key"]] = {
                "field": target["display"],
                "matched_header": best_header,
                "matched_by": best_candidate,
                "score": best_score,
                "values": values,
                "choices": choices,
                "row_values": row_values,
                "is_unique": _is_unique_row_values(row_values),
                "total_values": len(values),
            }

        return JSONResponse({
            "status": "success",
            "filename": filename,
            "matches": matches,
            "unmatched_keys": unmatched_keys,
            "matched_count": len(matches),
            "total_rows": len(rows),
            "message": f"Đã đối chiếu {len(matches)}/{len(targets)} trường"
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating reference field options: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi đọc file tham chiếu: {str(e)}")

@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload and parse Excel file"""
    try:
        # Get filename and extension
        filename = file.filename
        file_ext = os.path.splitext(filename)[1].lower()
        
        logger.info(f"Upload started: {filename} ({file_ext})")
        
        # Validate file type
        if file_ext not in ('.xlsx', '.xls'):
            raise HTTPException(
                status_code=400, 
                detail=f"Chi hỗ trợ file Excel (.xlsx, .xls). Bạn upload: {file_ext}"
            )
        
        # Read file content
        contents = await file.read()
        
        if not contents:
            raise HTTPException(
                status_code=400, 
                detail="File Excel rỗng hoặc không hợp lệ"
            )
        
        # Log file size
        logger.info(f"File size: {len(contents)} bytes")
        
        # For XLSX files: validate ZIP structure first
        if file_ext == '.xlsx':
            is_valid, error_msg = is_valid_xlsx(contents)
            if not is_valid:
                logger.error(f"XLSX validation failed: {error_msg}")
                raise HTTPException(
                    status_code=400,
                    detail=f"File Excel không hợp lệ: {error_msg}. Vui lòng kiểm tra file."
                )
        
        # Try to parse the Excel file
        excel_file = BytesIO(contents)
        error_messages = []
        headers = None
        rows = None
        
        # For .xlsx files: try openpyxl
        if file_ext == '.xlsx':
            try:
                logger.info("Attempting to parse .xlsx with openpyxl...")
                headers, rows, msg = parse_excel_with_openpyxl(excel_file)
            except Exception as e:
                error_msg = str(e)
                error_messages.append(f"openpyxl: {error_msg}")
                logger.error(f"openpyxl failed: {error_msg}")
                
                if "File contains no valid workbook part" in error_msg:
                    raise HTTPException(
                        status_code=400,
                        detail="File Excel bị hỏng hoặc không hợp lệ. Vui lòng:\n1. Mở file trong Excel\n2. Kiểm tra dữ liệu\n3. Lưu lại file\n4. Upload từ file mới"
                    )
                else:
                    raise HTTPException(status_code=400, detail=f"Không thể đọc file Excel: {error_msg}")
        
        # For .xls files: use xlrd
        elif file_ext == '.xls':
            try:
                logger.info("Attempting to parse .xls with xlrd...")
                headers, rows, msg = parse_xls_file(contents)
            except Exception as e:
                error_msg = str(e)
                logger.error(f"XLS parsing failed: {error_msg}")
                raise HTTPException(
                    status_code=400,
                    detail=f"Không thể đọc file .xls: {error_msg}"
                )
        
        # Store in memory
        if headers and rows:
            base_session_id = filename.replace('.xlsx', '').replace('.xls', '').replace(' ', '_')
            session_id = f"{base_session_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}"
            
            # Extract field metadata for better grouping
            field_metadata = extract_field_metadata(headers)
            
            excel_data_store[session_id] = {
                'user_id': current_user.id,
                'headers': headers,
                'rows': rows,
                'total_rows': len(rows),
                'filename': filename,
                'field_metadata': field_metadata,
                'created_at': datetime.utcnow().isoformat(),
                'history': []
            }

            # Ensure form + fields are available immediately after upload.
            form_id = _get_or_create_excel_form_id(db, current_user.id)
            _ensure_excel_fields(db, form_id, headers)
            
            logger.info(f"✓ Excel uploaded successfully: {len(rows)} rows, {len(headers)} columns (Format: {file_ext})")
            
            return JSONResponse({
                "status": "success",
                "session_id": session_id,
                "filename": filename,
                "headers": headers,
                "total_rows": len(rows),
                "field_metadata": field_metadata,
                "created_at": excel_data_store[session_id]['created_at'],
                "message": f"Tải file thành công! Tìm thấy {len(rows)} dòng dữ liệu"
            })
    
    except HTTPException:
        raise
    except Exception as e:
        error_detail = str(e)
        logger.error(f"Upload error: {error_detail}")
        raise HTTPException(
            status_code=500, 
            detail=f"Lỗi xử lý file: {error_detail}"
        )


@router.get("/data/{session_id}")
async def get_excel_data(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get Excel data for a session"""
    try:
        data = _get_session_data_or_404(session_id, current_user, db)
        return JSONResponse({
            "status": "success",
            "headers": data['headers'],
            "rows": data['rows'],
            "total_rows": data['total_rows'],
            "filename": data['filename'],
            "field_metadata": data.get('field_metadata', {})
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting Excel data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting data: {str(e)}")


@router.get("/row/{session_id}/{row_index}")
async def get_excel_row(
    session_id: str,
    row_index: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get specific row from Excel data"""
    try:
        data = _get_session_data_or_404(session_id, current_user, db)
        
        if row_index < 0 or row_index >= len(data['rows']):
            raise HTTPException(status_code=400, detail=f"Row index out of range. Valid range: 0-{len(data['rows'])-1}")
        
        row_data = data['rows'][row_index]
        
        return JSONResponse({
            "status": "success",
            "headers": data['headers'],
            "row_index": row_index,
            "row_data": row_data,
            "total_rows": data['total_rows'],
            "current_row": row_index + 1
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting row: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting row: {str(e)}")


@router.post("/add-page/{session_id}")
async def add_excel_page(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Append a new blank page (row) with the same headers in the current session."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)

        headers = session_data.get('headers', [])
        new_row = {
            str(header): ""
            for header in headers
            if str(header or "").strip()
        }

        rows = session_data.get('rows')
        if not isinstance(rows, list):
            rows = []
            session_data['rows'] = rows

        rows.append(new_row)
        session_data['total_rows'] = len(rows)

        _persist_session_to_db(
            db=db,
            session_id=session_id,
            user_id=current_user.id,
            filename=session_data.get('filename', ''),
            headers=headers,
            rows=rows,
            created_at=session_data.get('created_at') or datetime.utcnow().isoformat(),
            history=session_data.get('history', []),
        )

        return JSONResponse({
            "status": "success",
            "new_row_index": len(rows) - 1,
            "total_rows": len(rows),
            "row_data": new_row,
            "message": "Đã thêm trang mới"
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error adding excel page: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi thêm trang mới: {str(e)}")


@router.delete("/session/{session_id}")
async def delete_session(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a session"""
    try:
        _ = _get_session_data_or_404(session_id, current_user, db)

        db.query(ExcelTemplate).filter(
            ExcelTemplate.file_path == session_id,
            ExcelTemplate.template_name.like(f"{SESSION_TEMPLATE_PREFIX}%")
        ).delete(synchronize_session=False)
        db.commit()
        
        del excel_data_store[session_id]
        logger.info(f"Session deleted: {session_id}")
        
        return JSONResponse({
            "status": "success",
            "message": "Session deleted successfully"
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting session: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting session: {str(e)}")


@router.get("/sessions")
async def list_sessions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all active sessions"""
    try:
        query = db.query(ExcelTemplate).filter(
            ExcelTemplate.template_name.like(f"{SESSION_TEMPLATE_PREFIX}%")
        )

        if current_user.is_admin:
            rows = query.order_by(ExcelTemplate.created_at.desc()).all()
        else:
            rows = query.filter(ExcelTemplate.user_id == current_user.id).order_by(ExcelTemplate.created_at.desc()).all()

        sessions = []
        for row in rows:
            saved_count = 0
            if row.mapping_json:
                try:
                    payload = json.loads(row.mapping_json)
                    history = payload.get("history", []) if isinstance(payload, dict) else []
                    saved_count = len(_normalize_history_items(history))
                except Exception:
                    saved_count = 0

            sessions.append(
                {
                    "session_id": str(row.file_path),
                    "filename": str(row.original_filename or row.template_name.replace(SESSION_TEMPLATE_PREFIX, "")),
                    "total_rows": int(row.data_row_start or 0),
                    "created_at": row.created_at.isoformat() if row.created_at is not None else None,
                    "saved_count": saved_count,
                }
            )

        sessions.sort(key=lambda s: s.get('created_at') or '', reverse=True)
        
        return JSONResponse({
            "status": "success",
            "sessions": sessions,
            "total_sessions": len(sessions)
        })
    
    except Exception as e:
        logger.error(f"Error listing sessions: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listing sessions: {str(e)}")


@router.post("/save/{session_id}/{row_index}")
async def save_excel_row(
    session_id: str,
    row_index: int,
    payload: dict | None = Body(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Persist edited Excel row data into Entry history for cross-feature suggestions."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)
        if row_index < 0 or row_index >= len(session_data['rows']):
            raise HTTPException(status_code=400, detail="Row index out of range")

        row_data = payload.get("row_data") if isinstance(payload, dict) else None
        if not isinstance(row_data, dict):
            # Fallback: allow save using current row snapshot when client sends empty/invalid body.
            current_row = session_data['rows'][row_index]
            row_data = dict(current_row) if isinstance(current_row, dict) else {}

        headers = session_data.get('headers', [])
        form_id = _get_or_create_excel_form_id(db, current_user.id)

        # Update in-memory row for continuity in this Excel session.
        merged_row = dict(session_data['rows'][row_index])
        merged_row.update(row_data)
        session_data['rows'][row_index] = merged_row

        normalized_header_map: dict[str, str] = {
            _normalize_lookup_key(h): h for h in headers if (h or "").strip()
        }
        ho_header = None
        ten_header = None
        for normalized, original in normalized_header_map.items():
            if _is_family_name_header(normalized) and ho_header is None:
                ho_header = original
            elif _is_given_name_header(normalized) and ten_header is None:
                ten_header = original

        derived_headers = list(headers)
        if ho_header and ten_header and "Họ và tên" not in derived_headers:
            derived_headers.append("Họ và tên")

        field_map = _ensure_excel_fields(db, form_id, derived_headers)

        if ho_header and ten_header:
            ho_value = str(merged_row.get(ho_header, "") or "").strip()
            ten_value = str(merged_row.get(ten_header, "") or "").strip()
            full_name = " ".join([part for part in [ho_value, ten_value] if part]).strip()
            if full_name:
                merged_row["Họ và tên"] = full_name

        entries_to_insert: list[Entry] = []
        for header in derived_headers:
            field_name = (header or "").strip()
            if not field_name:
                continue

            value = str(merged_row.get(header, "") or "").strip()
            if not value:
                continue

            field_id = field_map.get(field_name)
            if not field_id:
                continue

            entries_to_insert.append(
                Entry(
                    user_id=current_user.id,
                    field_id=field_id,
                    form_id=form_id,
                    value=value
                )
            )

        if entries_to_insert:
            db.add_all(entries_to_insert)
            db.commit()

        history_item = _append_session_history(session_data, row_index, dict(merged_row))

        # Persist to history only after user explicitly clicks Save.
        _persist_session_to_db(
            db=db,
            session_id=session_id,
            user_id=current_user.id,
            filename=session_data.get('filename', ''),
            headers=session_data.get('headers', []),
            rows=session_data.get('rows', []),
            created_at=session_data.get('created_at') or datetime.utcnow().isoformat(),
            history=session_data.get('history', []),
        )

        return JSONResponse({
            "status": "success",
            "saved_entries": len(entries_to_insert),
            "history_item": history_item,
            "message": "Đã lưu lịch sử chỉnh sửa Excel"
        })
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error saving excel row history: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi lưu lịch sử Excel: {str(e)}")


@router.get("/history/{session_id}")
async def get_session_history(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return saved row history for an Excel session."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)
        history = _normalize_history_items(session_data.get("history"))
        session_data["history"] = history
        history_sorted = sorted(history, key=lambda item: item.get("saved_at", ""), reverse=True)
        return JSONResponse({
            "status": "success",
            "session_id": session_id,
            "history": history_sorted,
            "total": len(history_sorted),
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting session history: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi lấy lịch sử Excel: {str(e)}")


@router.put("/history/{session_id}/{history_id}")
async def update_session_history_item(
    session_id: str,
    history_id: str,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a history item and sync it back to session row data."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)
        history = _normalize_history_items(session_data.get("history"))
        row_data = payload.get("row_data") if isinstance(payload, dict) else None
        if not isinstance(row_data, dict):
            raise HTTPException(status_code=400, detail="row_data không hợp lệ")

        found_item = None
        for item in history:
            if item.get("history_id") == history_id:
                item["row_data"] = dict(row_data)
                item["saved_at"] = datetime.utcnow().isoformat()
                found_item = item
                break

        if not found_item:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi lịch sử")

        row_index = int(found_item.get("row_index", -1))
        if 0 <= row_index < len(session_data.get("rows", [])):
            session_data["rows"][row_index] = dict(row_data)
            session_data["total_rows"] = len(session_data["rows"])

        session_data["history"] = history
        _persist_session_to_db(
            db=db,
            session_id=session_id,
            user_id=session_data.get("user_id", current_user.id),
            filename=session_data.get("filename", ""),
            headers=session_data.get("headers", []),
            rows=session_data.get("rows", []),
            created_at=session_data.get("created_at") or datetime.utcnow().isoformat(),
            history=history,
        )

        return JSONResponse({
            "status": "success",
            "history_item": found_item,
            "message": "Đã cập nhật bản ghi lịch sử",
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating history item: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi cập nhật lịch sử Excel: {str(e)}")


@router.delete("/history/{session_id}/{history_id}")
async def delete_session_history_item(
    session_id: str,
    history_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete one history snapshot of a session."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)
        history = _normalize_history_items(session_data.get("history"))
        before_count = len(history)
        history = [item for item in history if item.get("history_id") != history_id]

        if len(history) == before_count:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi lịch sử")

        session_data["history"] = history
        _persist_session_to_db(
            db=db,
            session_id=session_id,
            user_id=session_data.get("user_id", current_user.id),
            filename=session_data.get("filename", ""),
            headers=session_data.get("headers", []),
            rows=session_data.get("rows", []),
            created_at=session_data.get("created_at") or datetime.utcnow().isoformat(),
            history=history,
        )

        return JSONResponse({
            "status": "success",
            "message": "Đã xóa bản ghi lịch sử",
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting history item: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi xóa lịch sử Excel: {str(e)}")


@router.get("/export/{session_id}")
async def export_session_data(
    session_id: str,
    export_format: str = Query("xlsx", alias="format"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export current Excel session data."""
    try:
        session_data = _get_session_data_or_404(session_id, current_user, db)
        headers = [str(h) for h in session_data.get("headers", []) if str(h).strip()]
        rows = session_data.get("rows", []) if isinstance(session_data.get("rows"), list) else []
        filename_root = os.path.splitext(str(session_data.get("filename", "excel_session")))[0]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        if not headers:
            raise HTTPException(status_code=400, detail="Session không có headers để xuất")

        fmt = (export_format or "xlsx").strip().lower()
        if fmt == "xlsx":
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Export"
            worksheet.append(headers)
            for row in rows:
                row_dict = row if isinstance(row, dict) else {}
                worksheet.append([row_dict.get(header, "") for header in headers])

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            export_name = f"{filename_root}_{timestamp}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{export_name}"'}
            )

        if fmt == "csv":
            output = BytesIO()
            text_buffer = output
            import io
            string_buffer = io.StringIO()
            writer = csv.writer(string_buffer)
            writer.writerow(headers)
            for row in rows:
                row_dict = row if isinstance(row, dict) else {}
                writer.writerow([row_dict.get(header, "") for header in headers])
            csv_text = string_buffer.getvalue()
            text_buffer.write(csv_text.encode("utf-8-sig"))
            text_buffer.seek(0)
            export_name = f"{filename_root}_{timestamp}.csv"
            return StreamingResponse(
                text_buffer,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{export_name}"'}
            )

        if fmt == "json":
            payload = {
                "session_id": session_id,
                "filename": session_data.get("filename"),
                "headers": headers,
                "rows": rows,
                "exported_at": datetime.utcnow().isoformat(),
            }
            content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
            output = BytesIO(content)
            export_name = f"{filename_root}_{timestamp}.json"
            return StreamingResponse(
                output,
                media_type="application/json; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{export_name}"'}
            )

        raise HTTPException(status_code=400, detail="Định dạng export không hợp lệ. Hỗ trợ: xlsx, csv, json")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting session: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi xuất dữ liệu Excel: {str(e)}")

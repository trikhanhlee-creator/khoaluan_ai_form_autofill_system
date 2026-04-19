#!/usr/bin/env python
"""
Excel Upload Diagnostics Tool
Checks all aspects of Excel upload functionality
"""
import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

import requests
import openpyxl
from io import BytesIO
from pathlib import Path
import time

class ExcelDiagnostics:
    def __init__(self):
        self.base_url = "http://127.0.0.1:8000"
        self.results = []
        self.backend_dir = Path(__file__).resolve().parents[2]
    
    def log(self, status, message):
        """Log a test result"""
        symbol = "✓" if status else "✗"
        color = "\033[92m" if status else "\033[91m"
        reset = "\033[0m"
        self.results.append((status, message))
        print(f"{color}{symbol}{reset} {message}")
    
    def print_header(self, title):
        """Print section header"""
        print(f"\n{'='*70}")
        print(f"  {title}")
        print(f"{'='*70}")
    
    def check_server_health(self):
        """Check if server is running"""
        self.print_header("1. Server Health Check")
        try:
            response = requests.get(f"{self.base_url}/health", timeout=5)
            if response.status_code == 200:
                self.log(True, "Server is running (HTTP 200)")
            else:
                self.log(False, f"Server returned HTTP {response.status_code}")
        except requests.ConnectionError:
            self.log(False, "Cannot connect to server - make sure it's running")
            self.log(False, f"  Tried: {self.base_url}")
            return False
        except Exception as e:
            self.log(False, f"Server check failed: {e}")
            return False
        return True
    
    def check_excel_route(self):
        """Check if Excel API route exists"""
        self.print_header("2. Excel API Route Check")
        try:
            # Try to get sessions (should work even if empty)
            response = requests.get(f"{self.base_url}/api/excel/sessions", timeout=5)
            if response.status_code == 200:
                self.log(True, "Excel API route is accessible")
                data = response.json()
                self.log(True, f"  Found {data.get('total_sessions', 0)} active sessions")
                return True
            else:
                self.log(False, f"Excel API returned HTTP {response.status_code}")
                return False
        except Exception as e:
            self.log(False, f"Excel API check failed: {e}")
            return False
    
    def test_valid_file_upload(self):
        """Test uploading a valid Excel file"""
        self.print_header("3. Valid File Upload Test")
        
        # Create test file
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Ho Ten'
            ws['B1'] = 'Email'
            ws['C1'] = 'Dien Thoai'
            ws['A2'] = 'Nguyen Van A'
            ws['B2'] = 'a@test.com'
            ws['C2'] = '0901234567'
            ws['A3'] = 'Tran Thi B'
            ws['B3'] = 'b@test.com'
            ws['C3'] = '0912345678'
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            self.log(True, f"Created test Excel file ({len(buffer.getvalue())} bytes)")
        except Exception as e:
            self.log(False, f"Failed to create test file: {e}")
            return False
        
        # Upload test file
        try:
            buffer.seek(0)
            files = {'file': ('test_valid.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{self.base_url}/api/excel/upload", files=files, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                self.log(True, f"Upload successful (HTTP 200)")
                self.log(True, f"  Session ID: {data.get('session_id')}")
                self.log(True, f"  Headers: {data.get('headers')}")
                self.log(True, f"  Data rows: {data.get('total_rows')}")
                return True
            else:
                self.log(False, f"Upload failed (HTTP {response.status_code})")
                self.log(False, f"  Response: {response.text}")
                return False
        except Exception as e:
            self.log(False, f"Upload test failed: {e}")
            return False
    
    def test_empty_file(self):
        """Test uploading an empty Excel file"""
        self.print_header("4. Empty File Error Handling Test")
        
        try:
            wb = openpyxl.Workbook()
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            files = {'file': ('empty.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{self.base_url}/api/excel/upload", files=files, timeout=10)
            
            if response.status_code == 400:
                data = response.json()
                self.log(True, "Empty file properly rejected (HTTP 400)")
                self.log(True, f"  Error message: {data.get('detail', 'N/A')}")
                return True
            else:
                self.log(False, f"Expected HTTP 400, got HTTP {response.status_code}")
                return False
        except Exception as e:
            self.log(False, f"Empty file test failed: {e}")
            return False
    
    def test_file_type_validation(self):
        """Test file type validation"""
        self.print_header("5. File Type Validation Test")
        
        invalid_files = [
            ('test.txt', 'text/plain', 'Text file'),
            ('test.pdf', 'application/pdf', 'PDF file'),
            ('test.csv', 'text/csv', 'CSV file'),
        ]
        
        all_passed = True
        for filename, mimetype, description in invalid_files:
            try:
                files = {'file': (filename, b'dummy content', mimetype)}
                response = requests.post(f"{self.base_url}/api/excel/upload", files=files, timeout=10)
                
                if response.status_code == 400:
                    self.log(True, f"{description} properly rejected")
                else:
                    self.log(False, f"{description} should be rejected but got HTTP {response.status_code}")
                    all_passed = False
            except Exception as e:
                self.log(False, f"File type validation test failed for {description}: {e}")
                all_passed = False
        
        return all_passed
    
    def test_sample_file(self):
        """Test with the sample file if it exists"""
        self.print_header("6. Sample File Upload Test")
        
        sample_path = self.backend_dir / "uploads" / "sample_data.xlsx"
        if not sample_path.exists():
            self.log(False, f"Sample file not found at {sample_path}")
            return False
        
        try:
            with open(sample_path, 'rb') as f:
                files = {'file': ('sample_data.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = requests.post(f"{self.base_url}/api/excel/upload", files=files, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                self.log(True, f"Sample file uploaded successfully")
                self.log(True, f"  Rows: {data.get('total_rows')}")
                self.log(True, f"  Headers: {len(data.get('headers', []))} columns")
                return True
            else:
                self.log(False, f"Sample file upload failed (HTTP {response.status_code})")
                return False
        except Exception as e:
            self.log(False, f"Sample file test failed: {e}")
            return False
    
    def check_static_files(self):
        """Check if static HTML files are accessible"""
        self.print_header("7. Static Files Check")
        
        pages = [
            ('/excel', 'Excel Upload Page'),
            ('/excel-upload.html', 'Excel Upload HTML'),
        ]
        
        all_passed = True
        for path, description in pages:
            try:
                response = requests.get(f"{self.base_url}{path}", timeout=5)
                if response.status_code == 200:
                    self.log(True, f"{description} is accessible")
                else:
                    self.log(False, f"{description} returned HTTP {response.status_code}")
                    all_passed = False
            except Exception as e:
                self.log(False, f"{description} check failed: {e}")
                all_passed = False
        
        return all_passed
    
    def print_summary(self):
        """Print test summary"""
        self.print_header("Summary")
        
        passed = sum(1 for status, _ in self.results if status)
        total = len(self.results)
        percentage = int((passed / total * 100)) if total > 0 else 0
        
        print(f"\nTotal Tests: {total}")
        print(f"Passed: {passed}")
        print(f"Failed: {total - passed}")
        print(f"Success Rate: {percentage}%")
        
        if passed == total:
            print("\n✓ All tests passed! Excel upload is working correctly.")
        elif passed >= total * 0.8:
            print("\n⚠ Most tests passed, but some issues detected.")
        else:
            print("\n✗ Multiple issues detected. Check the errors above.")
        
        print(f"{'='*70}\n")
    
    def run_diagnostics(self):
        """Run all diagnostics"""
        print("\n")
        print("╔" + "="*68 + "╗")
        print("║" + " "*68 + "║")
        print("║" + "  EXCEL UPLOAD DIAGNOSTICS TOOL".center(68) + "║")
        print("║" + " "*68 + "║")
        print("╚" + "="*68 + "╝")
        
        # Run all checks
        server_ok = self.check_server_health()
        if not server_ok:
            print("\n⚠ Cannot proceed without server. Start server first:")
            print("  cd backend")
            print("  python run.py")
            return
        
        self.check_excel_route()
        self.test_valid_file_upload()
        self.test_empty_file()
        self.test_file_type_validation()
        self.test_sample_file()
        self.check_static_files()
        
        self.print_summary()

if __name__ == "__main__":
    diagnostics = ExcelDiagnostics()
    diagnostics.run_diagnostics()

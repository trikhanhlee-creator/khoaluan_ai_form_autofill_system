#!/usr/bin/env python3
"""
Raw Excel structure inspection
"""
import openpyxl

file_path = r"c:\Users\KHANH\Downloads\demoexcel.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("\n" + "="*80)
print("RAW EXCEL STRUCTURE INSPECTION")
print("="*80)

print("\nFirst 3 rows - Raw cell values:")
for row_idx in range(1, 4):
    print(f"\nRow {row_idx}:")
    row_data = []
    for col_idx in range(1, 12):
        cell = ws.cell(row_idx, col_idx)
        value = cell.value
        is_merged = cell.coordinate in [mc.coord for mc in ws.merged_cells.ranges]
        status = " (MERGED)" if is_merged else ""
        row_data.append(f"Col{col_idx}: {repr(value)}{status}")
    for item in row_data:
        print(f"  {item}")

print("\n" + "="*80)
print("ANALYSIS:")
print("="*80)

row1_vals = [ws.cell(1, c).value for c in range(1, 10)]
row2_vals = [ws.cell(2, c).value for c in range(1, 10)]

print(f"\nRow 1 values: {row1_vals}")
print(f"Row 2 values: {row2_vals}")

# Check which row looks like headers
row1_non_empty = [x for x in row1_vals if x is not None]
row2_non_empty = [x for x in row2_vals if x is not None]

print(f"\nRow 1 non-empty cells: {len(row1_non_empty)}")
print(f"Row 2 non-empty cells: {len(row2_non_empty)}")

row1_text_count = sum(1 for x in row1_vals if isinstance(x, str))
row2_text_count = sum(1 for x in row2_vals if isinstance(x, str))

print(f"\nRow 1 text cells: {row1_text_count}/{len(row1_vals)}")
print(f"Row 2 text cells: {row2_text_count}/{len(row2_vals)}")

# Check for field keywords
field_keywords = ['mã', 'tên', 'họ', 'ngày', 'lớp', 'giới tính', 'sinh viên', 'đệm', 'chuyên cần', 'thường xuyên', 'quân sự', 'tình nguyện']

row1_keyword_count = 0
for val in row1_vals:
    if isinstance(val, str):
        val_lower = val.lower()
        for keyword in field_keywords:
            if keyword in val_lower:
                row1_keyword_count += 1
                break

row2_keyword_count = 0
for val in row2_vals:
    if isinstance(val, str):
        val_lower = val.lower()
        for keyword in field_keywords:
            if keyword in val_lower:
                row2_keyword_count += 1
                break

print(f"\nRow 1 cells with field keywords: {row1_keyword_count}")
print(f"Row 2 cells with field keywords: {row2_keyword_count}")

if row2_keyword_count > row1_keyword_count:
    print("\n✅ Row 2 should be used as the header row (more field keywords)")
else:
    print("\n⚠️  Row 1 is currently selected, but Row 2 has the actual field names")

print("\n" + "="*80 + "\n")

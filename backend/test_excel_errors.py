"""
Test to reproduce and fix the Excel upload error
"""
import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

import requests
import zipfile
from io import BytesIO

def test_corrupted_excel():
    """Test with corrupted Excel file"""
    print("\n" + "="*70)
    print("TEST 1: Corrupted/Invalid XLSX File")
    print("="*70)
    
    # Create a fake XLSX file that's invalid
    # It's a ZIP but missing required XML files
    
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w') as zf:
        zf.writestr('dummy.txt', 'This is not a valid XLSX')
    buffer.seek(0)
    
    files = {'file': ('corrupted.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
    
    print(f"Status: {response.status_code}")
    print(f"Response: {response.json()}")
    print()

def test_not_zip_file():
    """Test with file that's not a ZIP"""
    print("="*70)
    print("TEST 2: File Not ZIP (Invalid XLSX)")
    print("="*70)
    
    # Create a file that's clearly not a ZIP
    content = b"This is just text, not an Excel file"
    files = {'file': ('notanexcel.xlsx', content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
    
    print(f"Status: {response.status_code}")
    print(f"Response: {response.json()}")
    print()

def test_xls_file():
    """Test with old .xls format"""
    print("="*70)
    print("TEST 3: Old .xls Format")
    print("="*70)
    
    # Create a fake XLS file
    content = b"This is an old .xls format file"
    files = {'file': ('oldformat.xls', content, 'application/vnd.ms-excel')}
    response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
    
    print(f"Status: {response.status_code}")
    print(f"Response: {response.json()}")
    print()

def test_valid_file():
    """Test with valid Excel file"""
    print("="*70)
    print("TEST 4: Valid Excel File")
    print("="*70)
    
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Email'
    ws['A2'] = 'Test User'
    ws['B2'] = 'test@email.com'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    files = {'file': ('valid.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    response = requests.post('http://127.0.0.1:8000/api/excel/upload', files=files)
    
    print(f"Status: {response.status_code}")
    print(f"Response: {response.json()}")
    print()

if __name__ == "__main__":
    import time
    time.sleep(2)  # Wait for server
    
    try:
        test_corrupted_excel()
        test_not_zip_file()
        test_xls_file()
        test_valid_file()
    except Exception as e:
        print(f"\n✗ Error: {e}")
